# Import the faculty workload spreadsheet.
#   python manage.py import_workload_excel path/to/file.xlsx


import re

from datetime import date, timedelta

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from academics.models import (
    Faculty,
    Program,
    Cohort,
    AcademicPeriod,
    Lecturer,
    Course,
)

from accounts.models import (
    User,
)

from workload.models import Workload


# ==========================================================
# COLUMN LAYOUT
#
# One-based, matching the spreadsheet. Every cohort sheet uses
# the same layout; Cohort 5 has a trailing empty column, which
# changes nothing before column 12.
# ==========================================================

COL_GROUP = 1          # "SEM I" / row number
COL_PERIOD = 2
COL_LECTURER = 3
COL_QUALIFICATION = 4
COL_EMPLOYMENT = 5
COL_CODE = 6
COL_COURSE = 7
COL_CREDITS = 8
COL_LEVEL = 9
COL_DAYS = 10
COL_HOURS = 11
COL_STATUS = 12


SEMESTERS = [
    "Semester 1",
    "Semester 2",
    "Semester 3",
    "Semester 4",
]


# Titles stripped before two names are compared. "Dr. Eric
# Nizeyimana" and "Eric Nizeyimana" are one person.
TITLES = {
    "dr", "dr.", "prof", "prof.", "mr", "mr.",
    "mrs", "mrs.", "ms", "ms.", "miss", "rev", "rev.",
}


# The spreadsheet's day spellings.
DAY_FIXES = {
    "thusday": "Thursday",
    "thurday": "Thursday",
    "thursady": "Thursday",
    "tuesdady": "Tuesday",
    "sundy": "Sunday",
}

VALID_DAYS = [
    "Monday", "Tuesday", "Wednesday",
    "Thursday", "Friday", "Saturday", "Sunday",
]


# What "All days" means for Internship and Thesis.
#
# Taken literally it produced a seven-day week, which put
# Saturday and Sunday sittings on every internship and gave a
# thirteen-week placement 91 class dates. Nobody supervises on
# a weekend, so the phrase means the working week.
WORKING_DAYS = [
    "Monday", "Tuesday", "Wednesday",
    "Thursday", "Friday",
]


# ==========================================================
# SMALL HELPERS
# ==========================================================

def clean(value):
    """Trim a cell and collapse its internal whitespace."""

    if value is None:
        return ""

    return re.sub(r"\s+", " ", str(value)).strip()


def name_key(name):
    """
    A comparison key for a person's name.

    Titles are dropped and the remaining words are sorted, so
    "Dr. Pacifique Nizeyimana" and "Nizeyimana Pacifique" -- the
    same lecturer, written both ways in the same workbook --
    produce the same key.
    """

    words = re.split(r"[^A-Za-z]+", clean(name).lower())

    words = [
        word for word in words
        if word and word not in TITLES
    ]

    return " ".join(sorted(set(words)))


def course_key(name):
    """Comparison key for a course name."""

    return re.sub(
        r"[^a-z0-9]+",
        " ",
        clean(name).lower()
    ).strip()


def split_people(cell):
    """
    Split a lecturer cell into individual people.

    The spreadsheet separates co-teachers with a slash or a
    comma: "Dr.Eric Nizeyimana/ Dr. Kumar Kundan",
    "Dr. Eric Nizeyimana, Mr. david Hagumuwumva".
    """

    text = clean(cell)

    if not text:
        return []

    parts = re.split(r"\s*[/,]\s*", text)

    return [part for part in (clean(p) for p in parts) if part]


def normalise_days(cell):
    """
    Turn a course-day cell into "Sunday,Thursday".

    Handles the four forms in the workbook: comma separated,
    space separated, upper case, and the recurring misspelling
    "Thusday". "All days" means the whole week.

    Returns (days_string, note) where note describes any repair
    that was made, so it can be reported.
    """

    text = clean(cell)

    if not text:
        return "", ""

    if "all day" in text.lower():

        return (
            ",".join(WORKING_DAYS),
            "\"All days\" read as the working week, "
            "Monday to Friday"
        )

    tokens = [
        token
        for token in re.split(r"[,\s]+", text)
        if token
    ]

    days = []

    repaired = []

    for token in tokens:

        lowered = token.lower().strip(".")

        if lowered in DAY_FIXES:

            fixed = DAY_FIXES[lowered]

            repaired.append(f"{token} -> {fixed}")

        elif lowered.capitalize() in VALID_DAYS:

            fixed = lowered.capitalize()

        else:
            repaired.append(f"{token} not recognised as a day")
            continue

        if fixed not in days:
            days.append(fixed)

    note = "; ".join(repaired)

    return ",".join(days), note


# ==========================================================
# DATES
# ==========================================================

DATE_PATTERN = re.compile(
    r"^(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{2,4})$"
)


def parse_one_date(text):
    """
    Parse a single "D/M/YYYY" cell half.

    Day-first, which is how the whole workbook is written:
    "12/7/2026 - 16/8/2026" is July to August, not December.
    Returns None if the half is unusable, which happens -- one
    cell reads "../7/2026" and another ends in the year "203".
    """

    text = clean(text)

    match = DATE_PATTERN.match(text)

    if not match:
        return None

    day, month, year = (int(part) for part in match.groups())

    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None

    # A year outside this range is a typing slip, not a date.
    if not (2000 <= year <= 2100):
        return None

    try:
        return date(year, month, day)

    except ValueError:
        return None


def parse_period_cell(value):
    """
    Read the "Period" column.

    Three things turn up in it:

      a real datetime          -> the start date, no end
      "9/8/2026 - 13/9/2026"   -> a start and an end
      something broken         -> repair what can be repaired

    Returns (start, end, notes). Both dates may be None.
    """

    notes = []

    if value is None:
        return None, None, ["the period cell is empty"]

    # openpyxl hands back a datetime for the cells Excel stored
    # as dates rather than text.
    if hasattr(value, "year") and not isinstance(value, str):

        start = value.date() if hasattr(value, "date") else value

        return start, None, notes

    text = clean(value)

    if not text:
        return None, None, ["the period cell is empty"]

    halves = [
        half for half in re.split(r"\s*-\s*", text)
        if half.strip()
    ]

    if len(halves) == 1:
        return parse_one_date(halves[0]), None, notes

    start = parse_one_date(halves[0])

    end = parse_one_date(halves[1])

    raw_start, raw_end = halves[0], halves[1]

    # ---- Repair a broken half from its sibling ----
    #
    # "8/9/2030 -13/10/203" has a year missing a digit. The
    # start of the same cell says 2030 and October follows
    # September, so the end is 2030. This is arithmetic on the
    # cell's own contents, not a guess about the calendar.

    if end is None and start is not None:

        rescued = _rescue_half(raw_end, start, after=True)

        if rescued:
            end = rescued
            notes.append(
                f"end date \"{raw_end}\" read as {end} "
                f"using the year from the start of the same cell"
            )
        else:
            notes.append(f"end date \"{raw_end}\" could not be read")

    if start is None and end is not None:

        rescued = _rescue_half(raw_start, end, after=False)

        if rescued:
            start = rescued
            notes.append(
                f"start date \"{raw_start}\" read as {start} "
                f"using the year from the end of the same cell"
            )
        else:
            notes.append(f"start date \"{raw_start}\" could not be read")

    # ---- A block that ends before it starts ----
    #
    # "31/1/2026 - 7/3/2027" sits in a run of 2027 dates. Moving
    # the start forward one year is the only reading that keeps
    # the block in order.

    if start and end and end < start:

        # Two readings are possible, and the shorter block is
        # the right one. "13/12/2026 - 13/3/2026" is a term
        # ending in March 2027, not one starting in December
        # 2027: moving the end forward gives three months,
        # moving the start forward gives fifteen.

        end_shifted = _shift_year(end, +1)

        start_shifted = _shift_year(start, +1)

        if end_shifted and end_shifted > start:

            notes.append(
                f"end date {end} moved to {end_shifted}: the "
                f"cell ended before it began, and the year is "
                f"the only part that can be wrong"
            )

            end = end_shifted

        elif start_shifted and start_shifted <= end:

            notes.append(
                f"start date {start} moved to {start_shifted}: "
                f"the cell ended on {end}, before it began"
            )

            start = start_shifted

        else:
            notes.append(
                f"the period runs backwards ({start} to {end}); "
                f"the end date has been ignored"
            )
            end = None

    return start, end, notes


def _shift_year(value, years):
    """Same day and month, a different year. None on 29 February."""

    try:
        return date(value.year + years, value.month, value.day)

    except ValueError:
        return None


def _rescue_half(raw, sibling, after):
    """
    Rebuild a date whose year is mistyped, using the year of the
    other half of the same cell.
    """

    match = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*\d+\s*$", raw)

    if not match:
        return None

    day, month = int(match.group(1)), int(match.group(2))

    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None

    for year in (sibling.year, sibling.year + (1 if after else -1)):

        try:
            candidate = date(year, month, day)

        except ValueError:
            continue

        if after and candidate >= sibling:
            return candidate

        if not after and candidate <= sibling:
            return candidate

    return None


# ==========================================================
# THE COMMAND
# ==========================================================

class Command(BaseCommand):

    help = (
        "Import the faculty teaching-load spreadsheet into the "
        "system. Safe to run more than once."
    )

    def add_arguments(self, parser):

        parser.add_argument(
            "path",
            help="Path to the .xlsx workbook"
        )

        parser.add_argument(
            "--dry-run",
            action="store_true",
            help=(
                "Parse and report, then roll back. The import "
                "runs for real inside a transaction and is "
                "undone at the end, so a dry run exercises "
                "exactly the code a real run would."
            )
        )

        parser.add_argument(
            "--create-accounts",
            action="store_true",
            help=(
                "Create logins for lecturers and cohorts that "
                "have none. Email is left blank on purpose -- "
                "the administrator fills in the real address."
            )
        )

    # ------------------------------------------------------

    def handle(self, *args, **options):

        try:
            from openpyxl import load_workbook

        except ImportError:
            raise CommandError(
                "openpyxl is required. Install it with:\n"
                "    pip install openpyxl"
            )

        path = options["path"]

        self.dry_run = options["dry_run"]

        self.create_accounts = options["create_accounts"]

        try:
            workbook = load_workbook(path, data_only=True)

        except FileNotFoundError:
            raise CommandError(f"No such file: {path}")

        # Counters and report lines.
        self.created = {}
        self.reused = {}
        self.notes = []
        self.problems = []
        self.skipped = []
        self.accounts_needing_email = []

        try:
            with transaction.atomic():

                self._import(workbook)

                if self.dry_run:
                    raise _DryRun()

        except _DryRun:
            self.stdout.write(self.style.WARNING(
                "\nDRY RUN -- nothing was written.\n"
            ))

        self._report()

    # ------------------------------------------------------
    # COUNTERS
    # ------------------------------------------------------

    def _count(self, bucket, label, record=None):
        """
        Record one object under a label.

        Counted as a set of identities rather than a running
        total, because the same faculty is resolved once per
        sheet and the same lecturer on nearly every one. A
        plain increment reported ten faculties where there is
        one.
        """

        identity = (
            record.pk
            if record is not None and getattr(record, "pk", None)
            else object()
        )

        bucket.setdefault(label, set()).add(identity)

    def _made(self, label, record=None):
        self._count(self.created, label, record)

    def _kept(self, label, record=None):
        self._count(self.reused, label, record)

    # ------------------------------------------------------
    # MAIN
    # ------------------------------------------------------

    def _import(self, workbook):

        # Caches, so a name seen on ten sheets is resolved once.
        self.lecturer_cache = {}
        self.course_cache = {}

        self._load_existing()

        # The "Lecturers" sheet carries fuller names than the
        # cohort sheets do -- "Dr. Fabrice Sibomana" where a
        # cohort sheet says only "Dr. Fabrice". Read it first so
        # the better name is the one that gets stored.
        self._read_lecturer_sheet(workbook)

        sheets = [
            name for name in workbook.sheetnames
            if re.match(r"^\s*cohort\s*\d+", name.strip(), re.I)
        ]

        if not sheets:
            raise CommandError(
                "No sheets named 'Cohort<n>' were found in this "
                "workbook."
            )

        for sheet_name in sheets:
            self._import_sheet(workbook[sheet_name], sheet_name)

        if self.create_accounts:
            self._make_missing_accounts()

    # ------------------------------------------------------
    # EXISTING RECORDS
    # ------------------------------------------------------

    def _load_existing(self):
        """
        Index what is already in the database.

        Matching against this is what makes a second run safe,
        and it is why the import reuses the ten lecturers and
        fifteen courses already entered by hand rather than
        creating a second set beside them.
        """

        self.existing_lecturers = {}

        for lecturer in Lecturer.objects.all():

            self.existing_lecturers.setdefault(
                name_key(lecturer.name),
                lecturer
            )

        self.existing_courses = {}

        self.used_codes = set()

        for course in Course.objects.all():

            self.existing_courses.setdefault(
                course_key(course.name),
                course
            )

            self.used_codes.add(course.code.strip().lower())

    # ------------------------------------------------------
    # LECTURERS
    # ------------------------------------------------------

    def _read_lecturer_sheet(self, workbook):
        """
        Pick up full names and employment status from the
        per-lecturer sheet.
        """

        sheet = None

        for name in workbook.sheetnames:
            if name.strip().lower().startswith("lecturer"):
                sheet = workbook[name]
                break

        if sheet is None:
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if len(row) < 3:
                continue

            raw_name = clean(row[1])

            employment = clean(row[2])

            if not raw_name or raw_name.lower() == "instructor":
                continue

            for person in split_people(raw_name):
                self._resolve_lecturer(
                    person,
                    employment=employment
                )

    def _resolve_lecturer(
        self,
        raw_name,
        qualification="",
        employment=""
    ):
        """
        Find or create the Lecturer record for one person.

        Three passes, narrowest first:

          1. exact key match  -- "Dr. Eric Nizeyimana" == "Eric
             Nizeyimana"
          2. subset match     -- the database holds "Dr. Fabrice"
             and the sheet says "Dr. Fabrice Sibomana"; the
             shorter name is the same person, and the record is
             upgraded to the fuller one
          3. create

        A subset match is only accepted when exactly one record
        fits. Two candidates means the name is genuinely
        ambiguous, and guessing would attach a module to the
        wrong person's dashboard.
        """

        raw_name = clean(raw_name)

        if not raw_name:
            return None

        key = name_key(raw_name)

        if not key:
            return None

        if key in self.lecturer_cache:
            lecturer = self.lecturer_cache[key]
            self._enrich_lecturer(lecturer, qualification, employment)
            return lecturer

        # ---- 1. exact ----
        lecturer = self.existing_lecturers.get(key)

        # ---- 2. subset ----
        if lecturer is None:

            words = set(key.split())

            candidates = [
                (existing_key, record)
                for existing_key, record in self.existing_lecturers.items()
                if existing_key and (
                    set(existing_key.split()) <= words
                    or words <= set(existing_key.split())
                )
            ]

            if len(candidates) == 1:

                matched_key, lecturer = candidates[0]

                held_two_people = bool(re.search(r"[,/]", lecturer.name))

                if len(words) > len(set(matched_key.split())):

                    # The spreadsheet gives the fuller name:
                    # "Dr. Fabrice" in the database is
                    # "Dr. Fabrice Sibomana" here.
                    self.notes.append(
                        f"Lecturer \"{lecturer.name}\" renamed to "
                        f"\"{raw_name}\" -- the spreadsheet gives "
                        f"the full name for the same person"
                    )

                    lecturer.name = raw_name
                    lecturer.save(update_fields=["name"])

                elif held_two_people and not re.search(r"[,/]", raw_name):

                    # The record was created from a cell naming
                    # two co-teachers, so it is called
                    # "Nizeyimana Pacifique, Temitope" while its
                    # login belongs to one of them. Now that the
                    # cell has been split, the record stands for
                    # that one person and must be named for
                    # them -- otherwise their notification email
                    # is addressed to both.
                    self.notes.append(
                        f"Lecturer \"{lecturer.name}\" renamed to "
                        f"\"{raw_name}\" -- the record named two "
                        f"people, and now holds one. The other is "
                        f"a separate record. Please check the "
                        f"login attached to it is the right one."
                    )

                    lecturer.name = raw_name
                    lecturer.save(update_fields=["name"])

                self.existing_lecturers[key] = lecturer

            elif len(candidates) > 1:

                self.problems.append(
                    f"Lecturer \"{raw_name}\" matches more than one "
                    f"existing record ("
                    + ", ".join(r.name for _, r in candidates)
                    + "). A new record was created; please merge "
                    "them by hand."
                )

        # ---- 3. create ----
        if lecturer is None:

            lecturer = Lecturer.objects.create(
                name=raw_name,
                qualification=qualification,
                employment_status=employment or "Unspecified",
            )

            self._made("lecturers", lecturer)

            self.existing_lecturers[key] = lecturer

        else:
            self._kept("lecturers", lecturer)

            self._enrich_lecturer(lecturer, qualification, employment)

        self.lecturer_cache[key] = lecturer

        return lecturer

    def _enrich_lecturer(self, lecturer, qualification, employment):
        """Fill in blanks without overwriting anything set."""

        changed = []

        if qualification and not clean(lecturer.qualification):
            lecturer.qualification = qualification
            changed.append("qualification")

        if employment and not clean(lecturer.employment_status):
            lecturer.employment_status = employment
            changed.append("employment_status")

        if changed and lecturer.pk:
            lecturer.save(update_fields=changed)

    # ------------------------------------------------------
    # COURSES
    # ------------------------------------------------------

    def _resolve_course(self, name, code, credits, level, program):
        """
        Find or create a Course.

        Matched on name, not code. The workbook uses "MSDA 9233"
        for both Advanced database systems and Statistical
        Computing, and Course.code is unique, so trusting the
        code would collapse two courses into one. Where a code
        is already taken by a different course, a suffixed code
        is minted and the clash is reported.
        """

        name = clean(name)

        if not name:
            return None

        key = course_key(name)

        if key in self.course_cache:
            return self.course_cache[key]

        course = self.existing_courses.get(key)

        if course is not None:

            self._kept("courses", course)

            self.course_cache[key] = course

            return course

        code = clean(code)

        if not code:

            # Some rows name a module but give no code.
            initials = "".join(
                word[0] for word in re.findall(r"[A-Za-z]+", name)
            )[:4].upper() or "CRS"

            code = f"{initials}-{abs(hash(key)) % 9000 + 1000}"

            self.notes.append(
                f"Course \"{name}\" had no code in the "
                f"spreadsheet; \"{code}\" was generated"
            )

        final_code = code

        suffix = 2

        while final_code.strip().lower() in self.used_codes:

            final_code = f"{code}-{suffix}"

            suffix += 1

        if final_code != code:

            self.problems.append(
                f"Course code \"{code}\" is already used by a "
                f"different course, so \"{name}\" was given "
                f"\"{final_code}\". Please confirm the correct code."
            )

        course = Course.objects.create(
            program=program,
            code=final_code,
            name=name,
            credits=credits,
            level=level or "MSC",
        )

        self._made("courses", course)

        self.used_codes.add(final_code.strip().lower())

        self.existing_courses[key] = course

        self.course_cache[key] = course

        return course

    # ------------------------------------------------------
    # ONE COHORT SHEET
    # ------------------------------------------------------

    def _import_sheet(self, sheet, sheet_name):

        rows = list(sheet.iter_rows(values_only=True))

        def cell(row, column):
            index = column - 1
            return row[index] if index < len(row) else None

        # ---- Faculty, programme, cohort ----

        faculty_name = ""
        program_name = ""
        year_label = ""

        for row in rows[:6]:

            for value in row:

                text = clean(value)

                if not text:
                    continue

                if text.lower().startswith("faculty") and not faculty_name:
                    faculty_name = text

                elif text.lower().startswith("master") and not program_name:

                    # Sheets 5 to 9 read "...Big Data Analytics
                    # third semester". That trailing phrase is a
                    # leftover from copying the sheet, not part
                    # of the programme's name, and leaving it in
                    # creates a second programme that every one
                    # of those cohorts then hangs off.
                    program_name = re.sub(
                        r"\s*(first|second|third|fourth|1st|2nd|3rd|4th)?"
                        r"\s*semester\s*$",
                        "",
                        text,
                        flags=re.I
                    ).strip()

                elif "teaching load" in text.lower() and not year_label:
                    year_label = re.sub(
                        r"^\s*teaching load\s*",
                        "",
                        text,
                        flags=re.I
                    ).strip()

        faculty = self._get_or_create_named(
            Faculty,
            "faculties",
            faculty_name or "Faculty of Information Technology",
        )

        program = self._get_or_create_named(
            Program,
            "programmes",
            program_name or "Masters of science in Big Data Analytics",
            faculty=faculty,
        )

        number = re.search(r"\d+", sheet_name)

        cohort_name = f"Cohort {number.group()}" if number else sheet_name

        cohort = self._get_or_create(
            Cohort,
            "cohorts",
            program=program,
            name=cohort_name,
            defaults={"intake_year": self._intake_year(year_label)},
        )

        # ---- Walk the rows ----

        semester_index = -1

        seen_header = False

        pending = []

        for row_number, row in enumerate(rows, start=1):

            group = clean(cell(row, COL_GROUP))

            lecturer_cell = clean(cell(row, COL_LECTURER))

            # Header row: "Lecturer's Name" sits in column 3.
            if lecturer_cell.lower().startswith("lecturer"):

                seen_header = True

                explicit = self._semester_from_label(group)

                semester_index = (
                    explicit
                    if explicit is not None
                    else semester_index + 1
                )

                continue

            if not seen_header:
                continue

            # "SEM IV" starts the final semester on its own row.
            explicit = self._semester_from_label(group)

            if explicit is not None:
                semester_index = explicit

            course_name = clean(cell(row, COL_COURSE))

            if not lecturer_cell or not course_name:
                continue

            pending.append((row_number, row, max(semester_index, 0)))

        # ---- Periods, one per semester of this cohort ----
        #
        # Dates are read first so each period can span the rows
        # it actually contains, rather than a hard-coded term.

        parsed_rows = []

        spans = {}

        for row_number, row, semester_index in pending:

            start, end, notes = parse_period_cell(cell(row, COL_PERIOD))

            where = f"{sheet_name} row {row_number}"

            for note in notes:
                self.notes.append(f"{where}: {note}")

            if start is None:

                self.skipped.append(
                    f"{where}: no usable start date "
                    f"(\"{clean(cell(row, COL_PERIOD))}\"), "
                    f"so \"{clean(cell(row, COL_COURSE))}\" was not imported"
                )

                continue

            parsed_rows.append(
                (row_number, row, semester_index, start, end)
            )

            low, high = spans.get(semester_index, (start, end or start))

            spans[semester_index] = (
                min(low, start),
                max(high, end or start),
            )

        periods = {}

        for semester_index, (low, high) in spans.items():

            periods[semester_index] = self._get_or_create(
                AcademicPeriod,
                "academic periods",
                cohort=cohort,
                academic_year=year_label or str(low.year),
                semester=SEMESTERS[min(semester_index, 3)],
                defaults={
                    "teaching_period": f"{low:%d %b %Y} - {high:%d %b %Y}",
                    "start_date": low,
                    "end_date": high,
                },
            )

        # ---- Workload rows ----

        for row_number, row, semester_index, start, end in parsed_rows:

            self._import_row(
                sheet_name=sheet_name,
                row_number=row_number,
                row=row,
                cell=cell,
                cohort=cohort,
                program=program,
                period=periods[semester_index],
                start=start,
                end=end,
            )

    # ------------------------------------------------------
    # ONE WORKLOAD ROW
    # ------------------------------------------------------

    def _import_row(
        self,
        sheet_name,
        row_number,
        row,
        cell,
        cohort,
        program,
        period,
        start,
        end,
    ):

        where = f"{sheet_name} row {row_number}"

        course_name = clean(cell(row, COL_COURSE))

        # ---- People ----

        people = split_people(cell(row, COL_LECTURER))

        qualification = clean(cell(row, COL_QUALIFICATION))

        employment = clean(cell(row, COL_EMPLOYMENT))

        lecturers = []

        for person in people:

            record = self._resolve_lecturer(
                person,
                qualification=qualification,
                employment=employment,
            )

            if record and record not in lecturers:
                lecturers.append(record)

        if not lecturers:

            self.skipped.append(
                f"{where}: no lecturer could be read, so "
                f"\"{course_name}\" was not imported"
            )

            return

        # ---- Credits and hours ----

        credits, credit_note = self._read_credits(
            cell(row, COL_CREDITS),
            cell(row, COL_HOURS),
            course_name,
            where,
        )

        if credit_note:
            self.problems.append(credit_note)

        course = self._resolve_course(
            name=course_name,
            code=cell(row, COL_CODE),
            credits=credits,
            level=clean(cell(row, COL_LEVEL)),
            program=program,
        )

        if course is None:
            return

        # ---- Days ----

        days, day_note = normalise_days(cell(row, COL_DAYS))

        if day_note:
            self.notes.append(f"{where}: {day_note}")

        if not days:

            self.problems.append(
                f"{where}: \"{course_name}\" has no teaching days "
                f"in the spreadsheet. It was imported without a "
                f"schedule and shows as Pending; set its days on "
                f"the assignment screen."
            )

        # ---- Duration ----
        #
        # Internship and Thesis run a whole term, which no
        # credit value implies. Where the sheet states an end
        # date, the block length is taken from it rather than
        # recomputed.

        duration_weeks = None

        if end and days:

            weeks = max(1, round((end - start).days / 7))

            standard = 5 if credits >= 4 else 4

            if abs(weeks - standard) > 1:

                duration_weeks = weeks

                self.notes.append(
                    f"{where}: \"{course_name}\" runs {weeks} weeks "
                    f"in the spreadsheet rather than the usual "
                    f"{standard}; the stated span was kept"
                )

        # ---- Write ----

        workload, created = Workload.objects.get_or_create(
            cohort=cohort,
            academic_period=period,
            lecturer=lecturers[0],
            course=course,
            defaults={
                "start_date": start,
                "course_days": days,
                "duration_weeks": duration_weeks,
            },
        )

        if not created:

            workload.start_date = start
            workload.course_days = days
            workload.duration_weeks = duration_weeks
            workload.save()

            self._kept("workload assignments", workload)

        else:
            self._made("workload assignments", workload)

        workload.co_lecturers.set(lecturers[1:])

        if len(lecturers) > 1:
            self._count(self.created, "co-taught modules", workload)

    # ------------------------------------------------------
    # CREDITS
    # ------------------------------------------------------

    def _read_credits(self, credits_cell, hours_cell, course_name, where):
        """
        Read the credit value, and say so when it is doubtful.

        Every ordinary row has hours = credits x 15. Internship
        and Thesis break that: their credit column holds 60 and
        20 with the hours column empty, which are hours, not
        credits. Those two are converted and reported rather
        than stored as 60- and 20-credit courses, which would
        put 900 teaching hours against one module.
        """

        raw = clean(credits_cell)

        if not raw:
            return 3, (
                f"{where}: \"{course_name}\" has no credit value; "
                f"3 was assumed. Please confirm."
            )

        try:
            value = int(float(raw))

        except ValueError:
            return 3, (
                f"{where}: credit value \"{raw}\" for "
                f"\"{course_name}\" could not be read; 3 was "
                f"assumed. Please confirm."
            )

        if value <= 10:
            return value, ""

        # Larger than any credit value: read as hours.
        hours = value

        derived = max(1, round(hours / 15))

        return derived, (
            f"{where}: \"{course_name}\" shows {hours} in the "
            f"credits column with hours blank. Read as {hours} "
            f"hours, i.e. {derived} credits "
            f"({derived * 15} hours recorded). Please confirm."
        )

    # ------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------

    def _semester_from_label(self, label):
        """Turn "SEM III" into the index 2."""

        match = re.match(
            r"^\s*sem(?:ester)?\s*([IVX]+|\d+)\s*$",
            label or "",
            re.I
        )

        if not match:
            return None

        token = match.group(1).upper()

        romans = {"I": 0, "II": 1, "III": 2, "IV": 3}

        if token in romans:
            return romans[token]

        if token.isdigit():
            return max(0, int(token) - 1)

        return None

    def _intake_year(self, label):
        """First four-digit year in "Teaching Load 2026-2028"."""

        match = re.search(r"(20\d{2})", label or "")

        return int(match.group(1)) if match else date.today().year

    def _get_or_create(self, model, label, defaults=None, **lookup):
        """get_or_create that also counts."""

        existing = model.objects.filter(**lookup).first()

        if existing:
            self._kept(label, existing)
            return existing

        record = model.objects.create(**lookup, **(defaults or {}))

        self._made(label, record)

        return record

    def _get_or_create_named(self, model, label, name, **extra):
        """
        Find a record by name, ignoring spacing and case.

        The database holds "Masters of science in Big  Data
        Analytics" with two spaces, and the workbook writes it
        with two, three or one depending on the sheet. An exact
        match creates a second programme, and every cohort then
        hangs off the duplicate instead of the real one. Names
        are therefore compared with whitespace collapsed.
        """

        target = re.sub(r"\s+", " ", clean(name)).lower()

        for record in model.objects.filter(**extra):

            if re.sub(r"\s+", " ", record.name).strip().lower() == target:
                self._kept(label, record)
                return record

        created = model.objects.create(name=clean(name), **extra)

        self._made(label, created)

        return created

    # ------------------------------------------------------
    # ACCOUNTS
    # ------------------------------------------------------

    def _make_missing_accounts(self):
        """
        Give every lecturer and every cohort a login.

        The email address is left blank on purpose. Inventing
        one would mean the system reports a notification as sent
        while it goes nowhere, which is the one outcome worth
        avoiding. The accounts are listed at the end so the
        administrator can fill in the real addresses.

        Passwords are left unusable for the same reason: an
        account nobody has set a password on cannot be signed
        into by anyone else either.
        """

        for lecturer in Lecturer.objects.all():

            if lecturer.user_id:
                continue

            username = self._free_username(lecturer.name)

            user = User.objects.create(
                username=username,
                role="lecturer",
                email="",
            )

            user.set_unusable_password()

            user.save()

            lecturer.user = user

            lecturer.save(update_fields=["user"])

            self._made("lecturer logins", user)

            self.accounts_needing_email.append(
                f"{username} (lecturer: {lecturer.name})"
            )

        for cohort in Cohort.objects.all():

            if cohort.representative_id:
                continue

            username = self._free_username(f"{cohort.name} rep")

            user = User.objects.create(
                username=username,
                role="representative",
                email="",
            )

            user.set_unusable_password()

            user.save()

            cohort.representative = user

            cohort.save(update_fields=["representative"])

            self._made("representative logins", user)

            self.accounts_needing_email.append(
                f"{username} (representative: {cohort.name})"
            )

    def _free_username(self, source):
        """A readable username that is not already taken."""

        base = re.sub(r"[^A-Za-z0-9]+", "", clean(source).title()) or "User"

        base = base[:derive_max_length()]

        candidate = base

        suffix = 2

        while User.objects.filter(username=candidate).exists():

            candidate = f"{base}{suffix}"

            suffix += 1

        return candidate

    # ------------------------------------------------------
    # REPORT
    # ------------------------------------------------------

    def _report(self):

        write = self.stdout.write

        write("")
        write(self.style.MIGRATE_HEADING("What was written"))

        if not self.created and not self.reused:
            write("  nothing")

        for label in sorted(set(self.created) | set(self.reused)):

            made = len(self.created.get(label, ()))

            kept = len(self.reused.get(label, set()) - self.created.get(label, set()))

            write(
                f"  {label:<24} "
                f"{made:>4} created   "
                f"{kept:>4} already present"
            )

        self._section(
            "Rows not imported",
            self.skipped,
            self.style.ERROR,
        )

        self._section(
            "Needs your confirmation",
            self.problems,
            self.style.WARNING,
        )

        self._section(
            "Repairs and assumptions",
            self.notes,
            self.style.NOTICE if hasattr(self.style, "NOTICE") else self.style.WARNING,
        )

        if self.accounts_needing_email:

            write("")
            write(self.style.WARNING(
                "Accounts created without an email address"
            ))
            write(
                "  These people cannot be notified until an address\n"
                "  is set on the Users screen, and a password is set\n"
                "  before they can sign in.\n"
            )

            for line in self.accounts_needing_email:
                write(f"  - {line}")

        write("")

        if self.skipped or self.problems:
            write(self.style.WARNING(
                "Import finished with items needing attention "
                "(listed above)."
            ))

        else:
            write(self.style.SUCCESS("Import finished cleanly."))

        write("")

    def _section(self, title, lines, style):

        if not lines:
            return

        self.stdout.write("")
        self.stdout.write(style(f"{title} ({len(lines)})"))

        for line in lines:
            self.stdout.write(f"  - {line}")


def derive_max_length():
    """Maximum length of the username column."""

    return User._meta.get_field("username").max_length or 150


class _DryRun(Exception):
    """Rolls the transaction back at the end of a dry run."""
